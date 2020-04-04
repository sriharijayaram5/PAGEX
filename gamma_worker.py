import os
import _thread
from scipy.interpolate import interp1d
from bs4 import BeautifulSoup
import requests
from openpyxl import load_workbook
from openpyxl import Workbook
from mendeleev import element
from numpy import loadtxt, log10, array
import numpy as np
import scipy
from scipy import constants,optimize
from scipy.interpolate import InterpolatedUnivariateSpline
from chempy import Substance
import matplotlib as mpl
import matplotlib.pyplot as plt
from matplotlib import style
import pandas as pd
from pandas import read_csv
import math
import time
import sqlalchemy
import eel
import json
from datetime import datetime
import sys
sys.setrecursionlimit(5000)

n = scipy.constants.N_A

e_range = [
    1.000000E+03,
    1.500000E+03,
    2.000000E+03,
    3.000000E+03,
    4.000000E+03,
    5.000000E+03,
    6.000000E+03,
    8.000000E+03,
    1.000000E+04,
    1.500000E+04,
    2.000000E+04,
    3.000000E+04,
    4.000000E+04,
    5.000000E+04,
    6.000000E+04,
    8.000000E+04,
    1.000000E+05,
    1.500000E+05,
    2.000000E+05,
    3.000000E+05,
    4.000000E+05,
    5.000000E+05,
    6.000000E+05,
    8.000000E+05,
    1.000000E+06,
    1.022000E+06,
    1.250000E+06,
    1.500000E+06,
    2.000000E+06,
    2.044000E+06,
    3.000000E+06,
    4.000000E+06,
    5.000000E+06,
    6.000000E+06,
    7.000000E+06,
    8.000000E+06,
    9.000000E+06,
    1.000000E+07,
    1.100000E+07,
    1.200000E+07,
    1.300000E+07,
    1.400000E+07,
    1.500000E+07,
    1.600000E+07,
    1.800000E+07,
    2.000000E+07,
    2.200000E+07,
    2.400000E+07,
    2.600000E+07,
    2.800000E+07,
    3.000000E+07,
    4.000000E+07,
    5.000000E+07,
    6.000000E+07,
    8.000000E+07,
    1.000000E+08,
    1.500000E+08,
    2.000000E+08,
    3.000000E+08,
    4.000000E+08,
    5.000000E+08,
    6.000000E+08,
    8.000000E+08,
    1.000000E+09,
    1.500000E+09,
    2.000000E+09,
    3.000000E+09,
    4.000000E+09,
    5.000000E+09,
    6.000000E+09,
    8.000000E+09,
    1.000000E+10,
    1.500000E+10,
    2.000000E+10,
    3.000000E+10,
    4.000000E+10,
    5.000000E+10,
    6.000000E+10,
    8.000000E+10,
    1.000000E+11] 

xray_e_range = [0.0015, 0.002, 0.003, 0.004, 0.005, 0.006, 0.008,
                0.01, 0.015, 0.02, 0.03, 0.04, 0.05, 0.06, 0.08, 0.1,
                0.15, 0.2, 0.3, 0.4, 0.5, 0.6, 0.8, 1.0, 1.25, 1.5, 2.0,
                3.0, 4.0, 5.0, 6.0, 8.0, 10.0, 15.0, 20.0]
gp_energy_range = [
     0.015, 0.02, 0.03, 0.04, 0.05, 0.06, 0.08, 0.1, 0.15, 0.2, 0.3, 0.4, 0.5, 0.6, 0.8, 1.0, 1.5, 2.0, 3.0, 4.0, 5.0, 6.0, 8.0, 10.0, 15.0
     ]
def to_precision(x,p):
    """
    returns a string representation of x formatted with a precision of p
    Based on the webkit javascript implementation taken from here:
    https://code.google.com/p/webkit-mirror/source/browse/JavaScriptCore/kjs/number_object.cpp
    """
    x = float(x)

    if x == 0.:
        return float("0." + "0"*(p-1))

    out = []

    if x < 0:
        out.append("-")
        x = -x

    e = int(math.log10(x))
    tens = math.pow(10, e - p + 1)
    n = math.floor(x/tens)

    if n < math.pow(10, p - 1):
        e = e -1
        tens = math.pow(10, e - p+1)
        n = math.floor(x / tens)

    if abs((n + 1.) * tens - x) <= abs(n * tens -x):
        n = n + 1

    if n >= math.pow(10,p):
        n = n / 10.
        e = e + 1


    m = "%.*g" % (p, n)

    if e < -2 or e >= p:
        out.append(m[0])
        if p > 1:
            out.append(".")
            out.extend(m[1:p])
        out.append('e')
        if e > 0:
            out.append("+")
        out.append(str(e))
    elif e == (p -1):
        out.append(m)
    elif e >= 0:
        out.append(m[:e+1])
        if e+1 < len(m):
            out.append(".")
            out.extend(m[e+1:])
    else:
        out.append("0.")
        out.extend(["0"]*-(e+1))
        out.append(m)

    return float("".join(out))

def file_namer(z):
    if z < 10:
        zin = '00' + str(z)
    elif z < 100:
        zin = '0' + str(z)
    return(zin)

def comp_input():
    global weight_frac_list
    global comp_0
    global comp_1
    global comp_2
    if frac_flag:
        temp = True
        while temp:
            #ans = prompt(questions1,style=style_1)
            za = comp_1
            za1 = str(za)
            weight_frac_list = comp_2.split()
            weight_frac_list[:] = [float(x) for x in weight_frac_list]
            if len(za1.split())==len(weight_frac_list):
                temp = False
            else:
                print('Check the input! Number of consituent elements and weight fractions do not match.\n')
    else:
        #za = prompt(questions, style=style_1)['phone']
        za = comp_0
    return za
def fetch_compound():
    """This function takes user input of the compound."""
    global dict_comp
    global sub_unicode_name
    global elements_list
    global e_range
    global formula_for_text
    global photon_abs
    photon_abs = {}
    global photon_e_comp
    photon_e_comp = {}
    global R
    R = {}
    global compound_z_list
    compound_z_list = []
    global elements_list
    elements_list = []
    global ta_z
    ta_z = []
    global dict_comp
    dict_comp = {}
    global dict_element_to_para
    dict_element_to_para = {}
    global dict_element_to_photon
    dict_element_to_photon = {}
    global myu_comp
    myu_comp = []
    global sub_unicode_name
    sub_unicode_name = ''
    global flag
    flag = 0
    global myu_w_energy
    myu_w_energy = {}
    global photon
    photon = []
    global photon_comp
    photon_comp = {}
    global formula_for_text
    formula_for_text = ''
    global zeff_log
    zeff_log = {}
    global R_comp
    R_comp = {}
    global zeq_R
    zeq_R = {}
    global zeff_Ratio
    zeff_Ratio = {}
    global mec_comp
    mec_comp = {}
    global frac_flag
    global weight_frac_list
    
    #frac_flag = not prompt(weight_or_form,style=style_1)['frac_flag']==chc[0]
    za = comp_input()
    try:
        z = float(za)
        total_attenuation(file_namer(z))
        elements_list = element(z).symbol
        dict_element_to_para[tuple(elements_list)] = ta_z
        dict_comp[z] = 1
        

    except ValueError:
        if len(za) == 1:
            z = element(za.capitalize()).atomic_number
            total_attenuation(file_namer(z))
            elements_list = element(z).symbol
            dict_element_to_para[tuple(elements_list)] = ta_z
            compound_z_list = za.split()
            compound_z_list = [x.capitalize() for x in compound_z_list]
            za = ''.join(compound_z_list)
            sub = Substance.from_formula(za)
            sub_unicode_name = sub.unicode_name
            dict_comp = sub.composition
            
            formula_for_text = za
            
        else:
            
            compound_z_list = za.split()
            compound_z_list = [x.capitalize() for x in compound_z_list]
            za = ''.join(compound_z_list)
            sub = Substance.from_formula(za)
            sub_unicode_name = sub.unicode_name
            
            
            dict_comp = sub.composition
            formula_for_text = za

            for i,item in enumerate(compound_z_list):
                try:
                    item = float(item)
                    dict_comp[element(compound_z_list[i-1]).atomic_number] = item

                except ValueError:
                    z = element(item).atomic_number
                    total_attenuation(file_namer(z))
                    elements_list.append(item)
            

            elements_tuple = tuple(elements_list)
            dict_element_to_para[elements_tuple] = ta_z

def total_attenuation(zin):
    """Calculates total attenuation coefficients from NIST database."""
    list1 = []
    l = []
    zin1 = 'MDATX3.' + zin
    with open('NIST/' + zin1, 'r') as f:
        for line in f:
            l = line
            l = l.split()
            list1.extend(l)

    atomicn = list1[0]

    row_no = int(list1[3])

    if '1.00000E+03' in list1:
        start_index = list1.index('1.00000E+03')
    elif '1.000000E+03' in list1:
        start_index = list1.index('1.000000E+03')

    for i in range(start_index):
        list1.pop(0)

    for x in list1:
        try:
            list1[list1.index(x)] = float(x)
        except ValueError:
            list1.remove(x)
    #Calculation of partial and total interaction cross section (ta_wco)
    
    for i in e_range:
        index_energy = list1.index(i)
        energy = list1[index_energy]
        co_scat = list1[index_energy + row_no]
        inco_scat = list1[index_energy + 2 * row_no]
        photo_abs = list1[index_energy + 3 * row_no]
        pp_nf = list1[index_energy + 4 * row_no]
        pp_ef = list1[index_energy + 5 * row_no]
        ta_wco = co_scat + inco_scat + photo_abs + pp_nf + pp_ef
        m = element(int(atomicn)).mass
    #Total mass attenuation coefficients (ta_wco_1)
        ta_wco_1 = (ta_wco * n * 10**(-24)) / m
    #Partial mass attenuation coefficients
        ta_z.append([int(atomicn),
                     energy,
                     ta_wco_1,
                     (inco_scat * n * 10**(-24)) / m,
                     ((ta_wco - co_scat) * n * 10**(-24)) / m,
                     (co_scat * n * 10**(-24)) / m,
                     (inco_scat * n * 10**(-24)) / m,
                     (photo_abs * n * 10**(-24)) / m,
                     (pp_nf * n * 10**(-24)) / m,
                     (pp_ef * n * 10**(-24)) / m])

def myu():
    global myu_comp
    global elements_list
    global number_fraction
    global weight_fraction
    global weight_frac_list

    myu_comp = []

    w = []
    number_frac = []
    number_fraction = {}
    weight_fraction = {}
    global dict_comp
    elp = dict_element_to_para[tuple(elements_list)]
    #print(dict_comp)

    #Weight fraction - numerator calculation
    for i in elements_list:
        w.append(dict_comp[element(i).atomic_number] * element(i).mass)

    for i in elements_list:
        number_frac.append(dict_comp[element(i).atomic_number])
    #Weight fraction - denominator calculation
    x = 0.0
    z = 0.0
    total_number = 0
    for i in number_frac:
        total_number = total_number + i

    for j in elements_list:
        number_fraction[element(
            j).atomic_number] = number_frac[elements_list.index(j)] / total_number

    for i in w:
        x = x + i
    #Weight fraction - assignment to dictionary
    for j in elements_list:
        weight_fraction[element(
            j).atomic_number] = w[elements_list.index(j)] / x
    if frac_flag:
        for j in elements_list:
            weight_fraction[element(
                j).atomic_number] = weight_frac_list[elements_list.index(j)]
    


    denom = 0.0
    for j in elements_list:
        denom = denom + \
                (weight_fraction[element(j).atomic_number] / element(j).atomic_weight) 
            
    denom1 = denom * n
    k1 = int(len(elp) / len(elements_list))

    for k in range(k1):
        z = 0.0
        z1 = 0.0
        m_co = 0.0
        m_inco = 0.0
        m_pe = 0.0
        m_ppef = 0.0
        m_ppnf = 0.0
        for j in elements_list:
            z = z + (weight_fraction[element(j).atomic_number]
                     * elp[(elements_list.index(j) * k1) + k][2])
            z1 = z1 + (weight_fraction[element(j).atomic_number]
                     * elp[(elements_list.index(j) * k1) + k][4])
            m_co = m_co + (weight_fraction[element(j).atomic_number]
                     * elp[(elements_list.index(j) * k1) + k][5])
            m_inco = m_inco + (weight_fraction[element(j).atomic_number]
                     * elp[(elements_list.index(j) * k1) + k][6])
            m_pe = m_pe + (weight_fraction[element(j).atomic_number]
                     * elp[(elements_list.index(j) * k1) + k][7])
            m_ppef = m_ppef + (weight_fraction[element(j).atomic_number]
                     * elp[(elements_list.index(j) * k1) + k][8])
            m_ppnf = m_ppnf + (weight_fraction[element(j).atomic_number]
                     * elp[(elements_list.index(j) * k1) + k][9])
        if len(elements_list)==1:
            m = element(elements_list[0]).mass
            myu_comp.append([
                z,z1,z/denom1,z/denom1,
                m_co*m/n,
                m_inco*m/n,
                m_pe*m/n,
                m_ppef*m/n,
                m_ppnf*m/n,
                m_co,
                m_inco,
                m_pe,
                m_ppef,
                m_ppnf
                ])
        else:
            myu_comp.append([
                z,z1,z/denom1,z/denom1,
                m_co,
                m_inco,
                m_pe,
                m_ppef,
                m_ppnf
                ])

    for e in e_range:
        myu_w_energy[e] = myu_comp[e_range.index(e)]
    

def total_attenuation_1(zin):

    list1 = []
    l = []
    zin1 = 'MDATX3.' + zin
    global photon_abs
    photon_abs = {}
    global R
    R = {}
    with open('NIST/' + zin1, 'r') as f:
        for line in f:
            l = line
            l = l.split()
            list1.extend(l)
    atomicn = list1[0]
    
    row_no = int(list1[3])
    if '1.00000E+03' in list1:
        start_index = list1.index('1.00000E+03')
    elif '1.000000E+03' in list1:
        start_index = list1.index('1.000000E+03')

    for i in range(start_index):
        list1.pop(0)

    for x in list1:
        try:
            list1[list1.index(x)] = float(x)
        except ValueError:
            list1.remove(x)

    for i in e_range:
        index_energy = list1.index(i)
        energy = list1[index_energy]
        co_scat = list1[index_energy + row_no]
        inco_scat = list1[index_energy + 2 * row_no]
        photo_abs = list1[index_energy + 3 * row_no]
        pp_nf = list1[index_energy + 4 * row_no]
        pp_ef = list1[index_energy + 5 * row_no]
        ta_wco = co_scat + inco_scat + photo_abs + pp_nf + pp_ef

        m = element(int(atomicn)).mass
        ta_wco_1 = (ta_wco * n * 10**(-24)) / m

        # *10**(24) should also be there but gets cancelled, need it elsewhere for other calculations
        photon_abs[energy] = (ta_wco_1 * m / n)
        R[energy] = (inco_scat / (ta_wco - co_scat))


def zeff_by_log():
    global zeff_log
    global myu_comp
    global elements_list
    global photon
    global photon_comp
    global photon_e_comp
    global number_fraction
    global dict_comp
    
    photon_e_comp = {}
    photon = []
    elp = dict_element_to_para[tuple(elements_list)]
    photon_abs_element_list = []
    zeff_log = {}
    photon_abs_element_list = loadtxt(
        "element_photo_abs", usecols=(0), unpack=True)
    k1 = int(len(elp) / len(elements_list))
    
    
    denom = 0.0
    if frac_flag:
        for j in elements_list:
            denom = denom + \
                (weight_fraction[element(j).atomic_number] / element(j).atomic_weight)
        denom1 = denom * n

    for k in range(k1):
        z = 0.0
        for j in elements_list:
            z = z + (number_fraction[element(j).atomic_number] * elp[(
                elements_list.index(j) * k1) + k][2]) * element(j).atomic_weight

        photon.append(z / n)

    for e in e_range:
        photon_comp[e] = photon[e_range.index(e)]
    if frac_flag:
        for e in e_range:
            photon_comp[e] = myu_w_energy[e][0]/denom1   

    ele_photo_int_cross = []

    for i in range(98):
        a = list(photon_abs_element_list[i * len(e_range):(i + 1) * len(e_range)])
        ele_photo_int_cross.append(a)
    zno = np.arange(1,99)
    params = list(map(list,zip(*ele_photo_int_cross))) 
    z_comp = []
    z_comp[:] = [element(j).atomic_number for j in elements_list]

    zeff1={}
    for e in e_range:
        zeff1[e] = InterpolatedUnivariateSpline(zno, array(params[e_range.index(e)]) - photon_comp[e]).roots()

    avg = 0
    Aavg = 0
    z_total = 0
    #Mass weighted sum (Zi*Wi)- for finding best Zeff from multiple roots
    for j in z_comp:
        avg = avg +(j * weight_fraction[j])
        Aavg = Aavg + ( dict_comp[j] * element(j).mass )
        z_total = z_total + dict_comp[j]
    Aavg = Aavg/z_total
    sigma_a = {}
    for e in e_range:
        sigma_a[e] = photon_comp[e]

    for e in e_range:
        if len(zeff1[e])>1:
            zeff = min(zeff1[e], key=lambda x1:abs(x1-avg))
        else:
            zeff=zeff1[e][0]
        zeff_log[e] = [zeff,
                       1, 1,'NA','NA','NA',  (myu_w_energy[e][0]/sigma_a[e])*zeff]




def zeq_by_R(gp=False):
    global zeq_R
    global dict_comp
    global elements_list
    global R_comp
    global dict_comp
    global weight_fraction
    global B
    global BE
    global mfps
    global do_what
    global mean_free_path
    elp = dict_element_to_para[tuple(elements_list)]

    R_comp = {}
    R_comp_list = []
    R_comp = {}
    R_comp_list = []
    R_element_list = []
    R_this_element = []
    zeq_R = {}
    R_element_list = loadtxt("element_R", usecols=(0), unpack=True)
    k1 = int(len(elp) / len(elements_list))
    
    
    for k in range(k1):
        z = [0.0, 0.0]
        for j in elements_list:
            z[0] = z[0] + (weight_fraction[element(j).atomic_number]
                           * elp[(elements_list.index(j) * k1) + k][3])
            z[1] = z[1] + (weight_fraction[element(j).atomic_number]
                           * elp[(elements_list.index(j) * k1) + k][4])

        R_comp_list.append(z[0] / z[1])

    for e in e_range:
        R_comp[e] = R_comp_list[e_range.index(e)]

    ele_R = []

    for  i in range(98):
        R_this_element=R_element_list[i*len(e_range):(i+1)*len(e_range)] 
        ele_R.append(R_this_element)
    zno = np.arange(1,99)
    params = list(map(list,zip(*ele_R))) 
    z_comp = []
    z_comp[:] = [element(j).atomic_number for j in elements_list]

    zeff1={}
    for e in e_range:
        zeff1[e] = InterpolatedUnivariateSpline(zno, array(params[e_range.index(e)]) - R_comp[e]).roots()
        if len(elements_list)==1:
            zeff1[e] = [z_comp[0]]


    avg=0
    for j in z_comp:
        avg = avg +(j * weight_fraction[j])
    
    for e in e_range:
        if len(zeff1[e])>1:
            zeff = min(zeff1[e], key=lambda x1:abs(x1-avg))
        else:
            zeff=zeff1[e][0]
        
        zeq_R[e] = [zeff, math.floor(zeff), math.ceil(zeff), R_comp[e], 2, 2]
    
    
    prob_flag = True
    for e in gp_energy_range:
        if zeq_R[e*1000000][0]<4 or zeq_R[e*1000000][0]>82:
            if gp:
                print('\nData is not available for this element/compound! Zeq may be < 4 or > 82.\n')
                prob_flag = False
                return 'Back'
    if gp and prob_flag:
        b = get_gp('A',1)
        c = get_gp('A',2)
        a = get_gp('A',3)
        xk = get_gp('A',4)
        d = get_gp('A',5)

        b1 = get_gp('B',1)
        c1 = get_gp('B',2)
        a1 = get_gp('B',3)
        xk1 = get_gp('B',4)
        d1 = get_gp('B',5)
        
        B = {}
        BE = {}

        #mfp_q = prompt(user_mfp_list, style=style_1)['user_mfp_list']
        mfp_q = mean_free_path
        mfps = mfp_q.split()
        mfps[:] = [float(x) for x in mfps]
        for x in mfps:
            for i,e in enumerate(gp_energy_range):
                k1 = (
                    (c[i] * (x**a[i])) + d[i] *( (np.tanh( (x/xk[i])-2 ) - np.tanh( -2 )) / ( 1 - np.tanh( -2 ) ))
                )
                if mfps.index(x) == 0:
                    B[e*1000000] = [round(b[i],3), round(c[i],3), round(a[i],3), round(xk[i],3), round(d[i],3) ]
                if k1 == 1:
                    B[e*1000000].append(round(1 +(( b[i]-1 ) * x),3) )
                else:
                    B[e*1000000].append(round(1 +( ( (b[i] - 1) * ( k1**x - 1) ) / ( k1 - 1) ),3) )


            for i,e in enumerate(gp_energy_range):
                k1 = (
                    (c1[i] * (x**a1[i])) + d1[i] *( (np.tanh( (x/xk1[i])-2 ) - np.tanh( -2 )) / ( 1 - np.tanh( -2 ) ))
                )
                if mfps.index(x) == 0:
                    BE[e*1000000] = [ round(b1[i],3), round(c1[i],3), round(a1[i],3), round(xk1[i],3), round(d1[i],3) ]
                if k1 == 1:
                    BE[e*1000000].append(round(1 +(( b1[i]-1 ) * x),3) )
                else:
                    BE[e*1000000].append( round(1 + (( (b1[i] - 1) * ( k1**x - 1) ) / ( k1 - 1)),3) )
        return 'G-P fitting parameters and buildup factors - EABF, EBF'


def get_gp(db,param):
    global zeq_R
    good_z = []
    all_y=[]
    all_z=[]
    
    for  i in range(4,83):
        good_z.append(i)    
        zin = str(i)
        loc='ANSI_data/ANSI_'+db+'/DATA_'+zin
        try:
            a=read_csv(loc,delim_whitespace=True,header=None,usecols=[0,param],dtype=float)
            y1=list(array(a[param]))
            x=list(array(a[0]))
            y = []
            for e in gp_energy_range:
                y.append(y1[x.index(e)])
            all_y.append(y)
            all_z.append(i)
        except FileNotFoundError:
            continue
        

    a=list(zip(*all_y))
    inter_y=[]
    new_good_z = []
    for e in gp_energy_range:
        b = list(all_z)
        b.append(zeq_R[e*1000000][0])
        b.sort()
        new_good_z.append(b)

    new_param = []

    for i in range(len(a)):
        f = interp1d(all_z, a[i], kind='cubic')
        inter_y = f(new_good_z[i])
        new_param.append(inter_y[new_good_z[i].index(zeq_R[gp_energy_range[i]*1000000][0])])
    return new_param



def zeff_by_Ratio():
    '''Zeff by direct method'''
    global zeff_Ratio
    global elements_list
    global photon
    global photon_comp
    global photon_e_comp
    global number_fraction
    photon_e_comp = {}
    photon = []
    photon_e = []
    photon_e = []
    zeff_Ratio = {}
    global dict_comp
    elp = dict_element_to_para[tuple(elements_list)]
    k1 = int(len(elp) / len(elements_list))
    
    

    for k in range(k1):
        z = 0.0
        #To compute total molecular cross section
        for j in elements_list:
            z = z + (number_fraction[element(j).atomic_number] * elp[(
                elements_list.index(j) * k1) + k][2]) * element(j).atomic_weight

        photon.append(z / n)

    for e in e_range:
        photon_comp[e] = photon[e_range.index(e)]

    for k in range(k1):
        z = 0.0
        #To compute total electronic cross section
        for j in elements_list:
            z = z + (number_fraction[element(j).atomic_number] * elp[(elements_list.index(
                j) * k1) + k][2]) * element(j).atomic_weight / element(j).atomic_number

        photon_e.append(z / n)

    for e in e_range:
        photon_e_comp[e] = photon_e[e_range.index(e)]

    for e in e_range:
        zeff_Ratio[e] = [
            (photon_comp[e] / photon_e_comp[e]),
            photon_comp[e],
            photon_e_comp[e],'NA','NA','NA',
            myu_w_energy[e][0]/photon_e_comp[e]]
    

def stopping_power_compound_post():
    global elements_list
    global weight_fraction
    global density_mat
    
    formula = []

    for e in elements_list:
        #formula.extend(
        #    [e, str(weight_fraction[element(e).atomic_number]), '\n'])
        formula.extend(
            [e, str(round(weight_fraction[element(e).atomic_number],6)), '\n'])
    # formula=['H','0.11190674437968359','\n','O','0.8880932556203164']
    formula_req = ''
    for f in formula:
        formula_req = formula_req + f + ' '
    with requests.Session() as c:
        url = 'https://physics.nist.gov/cgi-bin/Star/estar-ut.pl'
        flag = 1
        while flag == 1:
            try:

                d = density_mat
                flag = 2

            except ValueError:
                print('Enter a number!')
        data1 = {'Name': 'Material',
                 'Density': d,
                 'Formulae': formula_req}
        a = True
        while a:
            try:
                r = c.post(url, data=data1)
                a = False
            except requests.exceptions.ConnectionError:
                eel.excel_alert("Please check internet conncection. This parameter requires an active connection.")

        soup = BeautifulSoup(r.text, 'html.parser')
        # print(soup.contents)
        results = soup.find('input').attrs['value']
        # print(results)
    with requests.Session() as c:
        url = 'https://physics.nist.gov/cgi-bin/Star/e_table-ut.pl'
        pairnum = -1
        for f in formula:
            try:
                float(f)
                pairnum = pairnum + 1
            except ValueError:
                pass

        lines = []
        for f in range(0, len(formula), 3):
            if formula[f] == '\n':
                continue
            lines.append(formula[f] + ' ' + formula[f + 1])
        data1 = {'I': results,
                 'ShowDefault': 'on',
                 'Name': 'Material',
                 'Density': d,
                 'pairnum': pairnum}
        for l in lines:
            data1['line' + str(lines.index(l))] = l
        a = True
        while a:
            try:
                r = c.post(url, data=data1)
                a = False
            except requests.exceptions.ConnectionError:
                eel.excel_alert("Please check internet conncection. This parameter requires an active connection.")
        soup = BeautifulSoup(r.text, 'html.parser')
        results = soup.find_all('pre')

        dat1 = results[0].contents[12:-1:2]
        dat1[:] = [x for x in dat1 if x != ' ']
        dat1[:] = [x for x in dat1 if x != '\x0c']
        dat1[:] = [x.strip() for x in dat1]
        with open('estar_comp', "w") as text_file:
            for d in dat1:
                print("{}".format(d), file=text_file)
        a, b, = loadtxt('estar_comp', usecols=(0, 3), unpack=True)

        return([a, b])
        


def zeff_electron_interaction():
    global elements_list
    global dict_comp
    global weight_fraction
    global zeff_ele
    global data
    global electron_msp
    electron_msp = {}
    zeff_ele = {}
    mass_stopping_power = []
    electron_int_cross = {}
    data = stopping_power_compound_post()
    mass_stopping_power = list(data[1])
    x = list(data[0])
    
    

    denom = 0.0

    for j in elements_list:

        denom = denom + \
            (weight_fraction[element(j).atomic_number] / element(j).atomic_weight)
    denom1 = denom * n
    for e in x:
        electron_int_cross[e] = mass_stopping_power[x.index(e)] / denom1
        electron_msp[e] = mass_stopping_power[x.index(e)]

    # Comparison begins
    ele_electron_int_cross = []

    for i in range(1, 99):
        z = element(i).atomic_number
        if z < 10:
            zin = '00' + str(z)
        else:
            zin = '0' + str(z)
        loc = 'EStar_data/DATA' + zin
        a=read_csv(loc,delim_whitespace=True,header=None,usecols=[0,3])
        y=array(a[3])
        b = n / element(i).atomic_weight
        y1 = y / b
        ele_electron_int_cross.append(list(y1))
    zno = np.arange(1,99)
    params = list(map(list,zip(*ele_electron_int_cross)))
    z_comp = []
    z_comp[:] = [element(j).atomic_number for j in elements_list]

    zeff1={}
    for e in x:
        solutions = InterpolatedUnivariateSpline(zno, params[x.index(e)] - electron_int_cross[e]).roots()
        zeff1[e] = solutions
        if len(elements_list)==1:
            zeff1[e] = [z_comp[0]]
    avg=0
    Aavg = 0
    z_total = 0
    for j in z_comp:
        avg = avg +(j * weight_fraction[j])
        Aavg = Aavg + ( dict_comp[j] * element(j).mass )
        z_total = z_total + dict_comp[j]
    Aavg = Aavg/z_total
    for e in x:
        if len(zeff1[e])>1:
            zeff = min(zeff1[e], key=lambda x1:abs(x1-avg))
        else:
            zeff=zeff1[e][0]
        
        zeff_ele[e] = [zeff,
            1, 1, electron_int_cross[e], 1, 1, electron_msp[e] / electron_int_cross[e]*zeff]
    

def zeff_proton_interaction():
    global elements_list
    global weight_fraction
    global zeff_proton
    global proton_energy_range
    global proton_msp
    proton_msp = {}
    zeff_proton = {}
    proton_int_cross = {}
    all_y = []
    all_z = []
    good_z = []
    msp = []
    msp_comp = []
    
    
    for i in range(1, 93):
        good_z.append(i)
        if i < 10:
            zin = '00' + str(i)
        else:
            zin = '0' + str(i)
        loc = 'PStar_data/DATA' + zin
        try:
            a=read_csv(loc,delim_whitespace=True,header=None,usecols=[3])
            y=array(a[3])
        except OSError:
            continue
        all_y.append(y)
        all_z.append(i)

    a = list(zip(*all_y))
    inter_y = []
    for i in range(len(a)):
        f = interp1d(all_z, a[i], kind='cubic')
        inter_y.append(f(good_z))

    all_new_y = list(zip(*inter_y))

    for i in range(1, 93):
        if i < 10:
            zin = '00' + str(i)
        else:
            zin = '0' + str(i)
        loc = 'PStar_data/DATA' + zin
        try:
            a=read_csv(loc,delim_whitespace=True,header=None,usecols=[0,3])
            y=array(a[3])
            x=array(a[0])
        except OSError:
            y = all_new_y[i - 1]
        msp.append(list(y))
    # Find comp MSP here
    for e in range(len(x)):
        z = 0.0
        for j in elements_list:
            z = z + (weight_fraction[element(j).atomic_number]
                     * msp[element(j).atomic_number - 1][e])

        msp_comp.append(z)
    
    denom = 0.0

    for j in elements_list:

        denom = denom + \
            (weight_fraction[element(j).atomic_number] / element(j).atomic_weight)
    denom1 = denom * n
    x = list(x)
    proton_energy_range=x
    for e in x:
        proton_int_cross[e] = msp_comp[x.index(e)] / denom1
        proton_msp[e] = msp_comp[x.index(e)]

    # Comparison begins
    ele_proton_int_cross = []

    for i in range(1, 93):
        z = element(i).atomic_number
        if z < 10:
            zin = '00' + str(z)
        else:
            zin = '0' + str(z)
        loc = 'PStar_data/DATA' + zin
        try:

            a=read_csv(loc,delim_whitespace=True,header=None,usecols=[3])
            y=array(a[3])
        except OSError:
            y = array(all_new_y[i - 1])
        b = n / element(i).atomic_weight
        y1 = y / b
        ele_proton_int_cross.append(list(y1))
    zno = np.arange(1,93)
    params = list(map(list,zip(*ele_proton_int_cross)))
    z_comp = []
    z_comp[:] = [element(j).atomic_number for j in elements_list]

    zeff1={}
    for e in x:
        solutions = InterpolatedUnivariateSpline(zno, params[x.index(e)] - proton_int_cross[e]).roots()
        zeff1[e] = solutions
        if len(elements_list)==1:
            zeff1[e] = [z_comp[0]]
    avg = 0
    Aavg = 0
    z_total = 0
    for j in z_comp:
        avg = avg +(j * weight_fraction[j])
        Aavg = Aavg + ( dict_comp[j] * element(j).mass )
        z_total = z_total + dict_comp[j]
    Aavg = Aavg/z_total

    for e in x:
        if len(zeff1[e])>1:
            zeff = min(zeff1[e], key=lambda x1:abs(x1-avg))
        else:
            zeff=zeff1[e][0]
        
        zeff_proton[e] = [zeff,
            1, 1, proton_int_cross[e], 1, 1, proton_msp[e] / proton_int_cross[e]*zeff]
    

def zeff_alpha_interaction():
    global elements_list
    global weight_fraction
    global zeff_alpha
    global alpha_energy_range
    global alpha_msp
    alpha_msp = {}
    zeff_alpha = {}

    alpha_int_cross = {}

    all_y = []
    all_z = []
    good_z = []
    msp = []
    msp_comp = []
    
    
    for i in range(1, 93):
        good_z.append(i)
        if i < 10:
            zin = '00' + str(i)
        else:
            zin = '0' + str(i)
        loc = 'AStar_data/DATA' + zin
        try:
            a=read_csv(loc,delim_whitespace=True,header=None,usecols=[0,3])
            x=array(a[0])
            y=array(a[3])
        except OSError:
            continue
        all_y.append(y)
        all_z.append(i)

    a = list(zip(*all_y))
    inter_y = []
    for i in range(len(a)):
        f = interp1d(all_z, a[i], kind='cubic')
        inter_y.append(f(good_z))

    all_new_y = list(zip(*inter_y))

    for i in range(1, 93):
        if i < 10:
            zin = '00' + str(i)
        else:
            zin = '0' + str(i)
        loc = 'AStar_data/DATA' + zin
        try:
            a=read_csv(loc,delim_whitespace=True,header=None,usecols=[0,3])
            x=array(a[0])
            y=array(a[3])
        except OSError:
            y = all_new_y[i - 1]
        msp.append(list(y))
    # Find comp MSP here
    for e in range(len(x)):
        z = 0.0
        for j in elements_list:
            z = z + (weight_fraction[element(j).atomic_number]
                     * msp[element(j).atomic_number - 1][e])

        msp_comp.append(z)
    

    denom = 0.0

    for j in elements_list:

        denom = denom + \
            (weight_fraction[element(j).atomic_number] / element(j).atomic_weight)
    denom1 = denom * n
    x = list(x)
    alpha_energy_range=x
    for e in x:
        alpha_int_cross[e] = msp_comp[x.index(e)] / denom1
        alpha_msp[e] = msp_comp[x.index(e)]

    ele_alpha_int_cross = []

    for i in range(1, 93):
        z = element(i).atomic_number
        if z < 10:
            zin = '00' + str(z)
        else:
            zin = '0' + str(z)
        loc = 'AStar_data/DATA' + zin
        try:
            a=read_csv(loc,delim_whitespace=True,header=None,usecols=[3])
            y=array(a[3])
        except OSError:
            y = array(all_new_y[i - 1])
        b = n / element(i).atomic_weight
        y1 = y / b
        ele_alpha_int_cross.append(list(y1))
    zno = np.arange(1,93)
    params = list(map(list,zip(*ele_alpha_int_cross)))
    z_comp = []
    z_comp[:] = [element(j).atomic_number for j in elements_list]

    zeff1={}
    
    for e in x:
        solutions = InterpolatedUnivariateSpline(zno, params[x.index(e)] - alpha_int_cross[e]).roots()
        zeff1[e] = solutions
        if len(elements_list)==1:
            zeff1[e] = [z_comp[0]]
    avg = 0
    Aavg = 0
    z_total = 0
    for j in z_comp:
        avg = avg +(j * weight_fraction[j])
        Aavg = Aavg + ( dict_comp[j] * element(j).mass )
        z_total = z_total + dict_comp[j]
    Aavg = Aavg/z_total
    for e in x:
        if len(zeff1[e])>1:
            zeff = min(zeff1[e], key=lambda x1:abs(x1-avg))
        else:
            zeff=zeff1[e][0]
        
        zeff_alpha[e] = [zeff,
            1,1, alpha_int_cross[e], 1,1, alpha_msp[e] / alpha_int_cross[e]*zeff]
    

def write_photon_abs_all_elements():  # only run to create initial file
    file = open("element_photo_abs", "w")
    for i in range(1, 100):
        #total_attenuation_2(i)
        # print(photon_abs)
        for e in e_range:
            file.write("%1.60f\n" % photon_abs[e])


def write_R_all_elements():  # only run to create initial file
    file = open("element_R", "w")
    for i in range(1, 100):
        #total_attenuation_2(i)
        # print(R)
        for e in e_range:
            file.write("%1.60f\n" % R[e])


def write_zeff_to_excel_sheet():
    global zeff_log
    
    sub = formula_for_text

    dest_filename = 'Excel_Sheets/Photon Zeff - Interpolation method.xlsx'
    if os.path.exists(dest_filename):
        wb = load_workbook(dest_filename)
        if frac_flag:
            sub = sub +'('
            for j in elements_list:
                sub = sub + str(round(weight_fraction[element(j).atomic_number],2)) + ','
            sub = sub.strip(',')
            sub = sub + ')'
            ws1 = wb.create_sheet(str(sub[:27]+'...'), 0)
        else:
            ws1 = wb.create_sheet(str(sub[:27]+'...'), 0)
    else:
        wb = Workbook()
        ws1 = wb.active
        if frac_flag:
            sub = sub +'('
            for j in elements_list:
                sub = sub + str(round(weight_fraction[element(j).atomic_number],2)) + ','
            sub = sub.strip(',')
            sub = sub + ')'
        ws1.title = str(sub[:27]+'...')
    ws1.cell(column=1, row=1, value='Energy (MeV)')
    ws1.cell(column=2, row=1, value='Zeff')
    
    if not False:
        ws1.cell(column=3, row=1, value='Neff (electrons/g)')
    mev=list(array(e_range)/1000000)

    for row in range(1, len(e_range) + 1):
        ws1.cell(column=1, row=row + 1, value=mev[row - 1])
        ws1.cell(column=2, row=row + 1, value=round(zeff_log[e_range[row - 1]][0],3))
        
        if not False:
            ws1.cell(column=3, row=row + 1, value=zeff_log[e_range[row - 1]][6])
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 15
    
    if not False:
        ws1.column_dimensions['C'].width = 25
    wb.save(filename=dest_filename)
    
    
    


def write_zeq_by_R_to_excel_sheet():
    global zeq_R
    
    sub = formula_for_text

    dest_filename = 'Excel_Sheets/Photon Zeq.xlsx'
    
    if os.path.exists(dest_filename):
        wb = load_workbook(dest_filename)
        if frac_flag:
            sub = sub +'('
            for j in elements_list:
                sub = sub + str(round(weight_fraction[element(j).atomic_number],2)) + ','
            sub = sub.strip(',')
            sub = sub + ')'
            ws1 = wb.create_sheet(str(sub[:27]+'...'), 0)
        else:
            ws1 = wb.create_sheet(str(sub[:27]+'...'), 0)
    else:
        wb = Workbook()
        ws1 = wb.active
        if frac_flag:
            sub = sub +'('
            for j in elements_list:
                sub = sub + str(round(weight_fraction[element(j).atomic_number],2)) + ','
            sub = sub.strip(',')
            sub = sub + ')'
        ws1.title = str(sub[:27]+'...')
    ws1.cell(column=1, row=1, value='Energy (MeV)')
    ws1.cell(column=2, row=1, value='Zeq')
    
    ws1.cell(column=3, row=1, value='R')
    
    mev=list(array(e_range)/1000000)
    for row in range(1, len(e_range) + 1):
        ws1.cell(column=1, row=row + 1, value=mev[row - 1])
        ws1.cell(column=2, row=row + 1, value=round(zeq_R[e_range[row - 1]][0],3))
        
        ws1.cell(column=3, row=row + 1, value=to_precision(zeq_R[e_range[row - 1]][3],4))
        
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 15
    
    ws1.column_dimensions['C'].width = 15
    
    wb.save(filename=dest_filename)
    
    
    

def write_BF_to_excel_sheet():
    global B
    global BE
    global mfps
    
    sub = formula_for_text

    dest_filename = 'Excel_Sheets/G-P fitting parameters and buildup factors.xlsx'
    
    if os.path.exists(dest_filename):
        wb = load_workbook(dest_filename)
        if frac_flag:
            sub = sub +'('
            for j in elements_list:
                sub = sub + str(round(weight_fraction[element(j).atomic_number],2)) + ','
            sub = sub.strip(',')
            sub = sub + ')'
            ws1 = wb.create_sheet(str(sub[:27]+'...'), 0)
        else:
            ws1 = wb.create_sheet(str(sub[:27]+'...'), 0)
    else:
        wb = Workbook()
        ws1 = wb.active
        if frac_flag:
            sub = sub +'('
            for j in elements_list:
                sub = sub + str(round(weight_fraction[element(j).atomic_number],2)) + ','
            sub = sub.strip(',')
            sub = sub + ')'
        ws1.title = str(sub[:27]+'...')
    ws1.cell(column=1, row=1, value='Energy (MeV)')
    ws1.cell(column=2, row=1, value='b')
    ws1.cell(column=3, row=1, value='c')
    ws1.cell(column=4, row=1, value='a')
    ws1.cell(column=5, row=1, value='Xk')
    ws1.cell(column=6, row=1, value='d')
    for i,x in enumerate(mfps):
        ws1.cell(column=6+i+1, row=1, value=f'EABF {x}')
    
    ws1.cell(column=6+len(mfps)+1, row=1, value='b')
    ws1.cell(column=7+len(mfps)+1, row=1, value='c')
    ws1.cell(column=8+len(mfps)+1, row=1, value='a')
    ws1.cell(column=9+len(mfps)+1, row=1, value='Xk')
    ws1.cell(column=10+len(mfps)+1, row=1, value='d')
    for i,x in enumerate(mfps):
        ws1.cell(column=10+len(mfps)+1+i+1, row=1, value=f'EBF {x}')

    mev=list(gp_energy_range)
    conv = 1000000
    for row in range(1, len(gp_energy_range) + 1):
        ws1.cell(column=1, row=row + 1, value=mev[row - 1])
        ws1.cell(column=2, row=row + 1, value=B[gp_energy_range[row - 1]*conv][0])
        ws1.cell(column=3, row=row + 1, value=B[gp_energy_range[row - 1]*conv][1])
        ws1.cell(column=4, row=row + 1, value=B[gp_energy_range[row - 1]*conv][2])
        ws1.cell(column=5, row=row + 1, value=B[gp_energy_range[row - 1]*conv][3])
        ws1.cell(column=6, row=row + 1, value=B[gp_energy_range[row - 1]*conv][4])
        for i,x in enumerate(mfps):
            ws1.cell(column=6+i+1, row=row + 1, value=B[gp_energy_range[row - 1]*conv][5+i])

        ws1.cell(column=6+len(mfps)+1, row=row + 1, value=BE[gp_energy_range[row - 1]*conv][0])
        ws1.cell(column=7+len(mfps)+1, row=row + 1, value=BE[gp_energy_range[row - 1]*conv][1])
        ws1.cell(column=8+len(mfps)+1, row=row + 1, value=BE[gp_energy_range[row - 1]*conv][2])
        ws1.cell(column=9+len(mfps)+1, row=row + 1, value=BE[gp_energy_range[row - 1]*conv][3])
        ws1.cell(column=10+len(mfps)+1, row=row + 1, value=BE[gp_energy_range[row - 1]*conv][4])
        for i,x in enumerate(mfps):
            ws1.cell(column=10+len(mfps)+1+i+1, row=row + 1, value=BE[gp_energy_range[row - 1]*conv][5+i])
    ws1.column_dimensions['A'].width = 15

    wb.save(filename=dest_filename)
    
    
    

def write_zeff_by_Ratio_to_excel_sheet():
    global zeff_Ratio
    global myu_w_energy
    
    sub = formula_for_text
    dest_filename = 'Excel_Sheets/Photon Zeff - Direct method.xlsx'
    if os.path.exists(dest_filename):
        wb = load_workbook(dest_filename)
        if frac_flag:
            sub = sub +'('
            for j in elements_list:
                sub = sub + str(round(weight_fraction[element(j).atomic_number],2)) + ','
            sub = sub.strip(',')
            sub = sub + ')'
            ws1 = wb.create_sheet(str(sub[:27]+'...'), 0)
        else:
            ws1 = wb.create_sheet(str(sub[:27]+'...'), 0)
    else:
        wb = Workbook()
        ws1 = wb.active
        if frac_flag:
            sub = sub +'('
            for j in elements_list:
                sub = sub + str(round(weight_fraction[element(j).atomic_number],2)) + ','
            sub = sub.strip(',')
            sub = sub + ')'
        ws1.title = str(sub[:27]+'...')
    ws1.cell(column=1, row=1, value='Energy (MeV)')
    ws1.cell(column=4, row=1, value='Zeff')
    ws1.cell(column=2, row=1, value='σₐ Average Cross Section per Atom (cm²/atom)')
    ws1.cell(column=3, row=1, value='σₑ Average Cross Section per Electron (cm²/electron)')
    ws1.cell(column=5, row=1, value='Neff (electrons/g)')
    mev=list(array(e_range)/1000000)
    for row in range(1, len(e_range) + 1):
        ws1.cell(column=1, row=row + 1, value=mev[row - 1])
        ws1.cell(column=4, row=row + 1, value=round(zeff_Ratio[e_range[row - 1]][0],3))
        ws1.cell(column=2, row=row + 1, value=to_precision(zeff_Ratio[e_range[row - 1]][1],4))
        ws1.cell(column=3, row=row + 1, value=to_precision(zeff_Ratio[e_range[row - 1]][2],4))
        ws1.cell(column=5, row=row + 1, value=to_precision(zeff_Ratio[e_range[row - 1]][6],4))
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 25
    ws1.column_dimensions['C'].width = 25
    ws1.column_dimensions['D'].width = 15
    ws1.column_dimensions['E'].width = 25
    
    wb.save(filename=dest_filename)
    
    
    


def write_zeff_ele_to_excel_sheet(x):
    global zeff_ele
    global electron_msp
    sub = formula_for_text

    dest_filename = 'Excel_Sheets/Electron interaction parameters.xlsx'
    if os.path.exists(dest_filename):
        wb = load_workbook(dest_filename)
        if frac_flag:
            sub = sub +'('
            for j in elements_list:
                sub = sub + str(round(weight_fraction[element(j).atomic_number],2)) + ','
            sub = sub.strip(',')
            sub = sub + ')'
            ws1 = wb.create_sheet(str(sub[:27]+'...'), 0)
        else:
            ws1 = wb.create_sheet(str(sub[:27]+'...'), 0)
    else:
        wb = Workbook()
        ws1 = wb.active
        if frac_flag:
            sub = sub +'('
            for j in elements_list:
                sub = sub + str(round(weight_fraction[element(j).atomic_number],2)) + ','
            sub = sub.strip(',')
            sub = sub + ')'
        ws1.title = str(sub[:27]+'...')
    ws1.cell(column=1, row=1, value='Energy (MeV)')
    if len(elements_list)>1:
        ws1.cell(column=4, row=1, value='Zeff')
    
    if len(elements_list)>1:
        ws1.cell(column=5, row=1, value='Neff (electrons/g)')
    ws1.cell(column=2, row=1, value='S(E)/ρ (MeV cm²/g)')
    ws1.cell(column=3, row=1, value='Sc (MeV cm²/atom)')
    for row in range(1, len(x) + 1):
        ws1.cell(column=1, row=row + 1, value=x[row - 1])
        if len(elements_list)>1:
            ws1.cell(column=4, row=row + 1, value=round(zeff_ele[x[row - 1]][0],3))
        
        if len(elements_list)>1:
            ws1.cell(column=5, row=row + 1, value=to_precision(zeff_ele[x[row - 1]][6],4))
        ws1.cell(column=2, row=row + 1, value=round(electron_msp[x[row - 1]],3))
        ws1.cell(column=3, row=row + 1, value=to_precision(zeff_ele[x[row - 1]][3],4))
    
 
    if not False:
        ws1.column_dimensions['E'].width = 25
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['D'].width = 15
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['E'].width = 15
    wb.save(filename=dest_filename)
    
    
    


def write_zeff_proton_to_excel_sheet(x):
    global zeff_proton
    global proton_msp
    sub = formula_for_text

    dest_filename = 'Excel_Sheets/Proton interaction parameters.xlsx'
    if os.path.exists(dest_filename):
        wb = load_workbook(dest_filename)
        if frac_flag:
            sub = sub +'('
            for j in elements_list:
                sub = sub + str(round(weight_fraction[element(j).atomic_number],2)) + ','
            sub = sub.strip(',')
            sub = sub + ')'
            ws1 = wb.create_sheet(str(sub[:27]+'...'), 0)
        else:
            ws1 = wb.create_sheet(str(sub[:27]+'...'), 0)
    else:
        wb = Workbook()
        ws1 = wb.active
        if frac_flag:
            sub = sub +'('
            for j in elements_list:
                sub = sub + str(round(weight_fraction[element(j).atomic_number],2)) + ','
            sub = sub.strip(',')
            sub = sub + ')'
        ws1.title = str(sub[:27]+'...')
    ws1.cell(column=1, row=1, value='Energy (MeV)')
    if len(elements_list)>1:
        ws1.cell(column=4, row=1, value='Zeff')

    if len(elements_list)>1:
        ws1.cell(column=5, row=1, value='Neff (electrons/g)')
    ws1.cell(column=2, row=1, value='S(E)/ρ (MeV cm²/g)')
    ws1.cell(column=3, row=1, value='Sc (MeV cm²/atom)')
    for row in range(1, len(x) + 1):
        ws1.cell(column=1, row=row + 1, value=x[row - 1])
        if len(elements_list)>1:
            ws1.cell(column=4, row=row + 1, value=round(zeff_proton[x[row - 1]][0],3))

        if len(elements_list)>1:
            ws1.cell(column=5, row=row + 1, value=to_precision(zeff_proton[x[row - 1]][6],4))
        ws1.cell(column=2, row=row + 1, value=round(proton_msp[x[row - 1]],3))
        ws1.cell(column=3, row=row + 1, value=to_precision(zeff_proton[x[row - 1]][3],4))
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['D'].width = 15

    if not False:
        ws1.column_dimensions['E'].width = 25
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['D'].width = 15
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['E'].width = 15
    wb.save(filename=dest_filename)
    
    
    


def write_zeff_alpha_to_excel_sheet(x):
    global zeff_alpha
    global alpha_msp
    
    sub = formula_for_text

    dest_filename = 'Excel_Sheets/Alpha particle interaction parameters.xlsx'
    if os.path.exists(dest_filename):
        wb = load_workbook(dest_filename)
        if frac_flag:
            sub = sub +'('
            for j in elements_list:
                sub = sub + str(round(weight_fraction[element(j).atomic_number],2)) + ','
            sub = sub.strip(',')
            sub = sub + ')'
            ws1 = wb.create_sheet(str(sub[:27]+'...'), 0)
        else:
            ws1 = wb.create_sheet(str(sub[:27]+'...'), 0)
    else:
        wb = Workbook()
        ws1 = wb.active
        if frac_flag:
            sub = sub +'('
            for j in elements_list:
                sub = sub + str(round(weight_fraction[element(j).atomic_number],2)) + ','
            sub = sub.strip(',')
            sub = sub + ')'
        ws1.title = str(sub[:27]+'...')
    ws1.cell(column=1, row=1, value='Energy (MeV)')
    if len(elements_list)>1:
        ws1.cell(column=4, row=1, value='Zeff')

    if len(elements_list)>1:
        ws1.cell(column=5, row=1, value='Neff (electrons/g)')
    ws1.cell(column=2, row=1, value='S(E)/ρ (MeV cm²/g)')
    ws1.cell(column=3, row=1, value='Sc (MeV cm²/atom)')
    
    for row in range(1, len(x) + 1):
        ws1.cell(column=1, row=row + 1, value=x[row - 1])
        if len(elements_list)>1:
            ws1.cell(column=4, row=row + 1, value=round(zeff_alpha[x[row - 1]][0],3))

        if len(elements_list)>1:
            ws1.cell(column=5, row=row + 1, value=to_precision(zeff_alpha[x[row - 1]][6],4))
        ws1.cell(column=2, row=row + 1, value=round(alpha_msp[x[row - 1]],3))
        ws1.cell(column=3, row=row + 1, value=to_precision(zeff_alpha[x[row - 1]][3],4))
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['D'].width = 15

    if not False:
        ws1.column_dimensions['E'].width = 25
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['D'].width = 15
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['E'].width = 15
    wb.save(filename=dest_filename)
    
    
    

def write_myu_to_excel_sheet():
    global myu_w_energy
    
    sub = formula_for_text
    dest_filename = 'Excel_Sheets/Photon mass attenuation and interaction cross section parameters.xlsx'
    if os.path.exists(dest_filename):
        wb = load_workbook(dest_filename)
        if frac_flag:
            sub = sub +'('
            for j in elements_list:
                sub = sub + str(round(weight_fraction[element(j).atomic_number],2)) + ','
            sub = sub.strip(',')
            sub = sub + ')'
            ws1 = wb.create_sheet(str(sub[:27]+'...'), 0)
        else:
            ws1 = wb.create_sheet(str(sub[:27]+'...'), 0)
    else:
        wb = Workbook()
        ws1 = wb.active
        if frac_flag:
            sub = sub +'('
            for j in elements_list:
                sub = sub + str(round(weight_fraction[element(j).atomic_number],2)) + ','
            sub = sub.strip(',')
            sub = sub + ')'
        ws1.title = str(sub[:27]+'...')
    output_choices1 = [

        'Partial/total interaction cross sections',
        'Partial/total mass attenuation coefficients',
        'Write both'
    ]
    output_questions1 = {

        'type': 'list',
        'name': 'output_type',
        'message': 'Choose output: ',
        'choices': output_choices1
    }
    sigma_flag = False           
    both_flag = False
    if len(elements_list)==1:
        #op = prompt(output_questions1,style=style_1)['output_type']        
        op = "Write both"
        if op == output_choices1[2]:
            both_flag = True
        elif op == output_choices1[0]:
            sigma_flag = True


    ws1.cell(column=1, row=1, value='Energy (MeV)')
    if both_flag:
        ws1.cell(column=2, row=1, value='σ(coh)  (cm²/atom)')
        ws1.cell(column=3, row=1, value='σ(incoh)  (cm²/atom)')
        ws1.cell(column=4, row=1, value='σ(pe)  (cm²/atom)')
        ws1.cell(column=5, row=1, value='σ(pair)  (cm²/atom)')
        ws1.cell(column=6, row=1, value='σ(trip)  (cm²/atom)')

        ws1.cell(column=7, row=1, value='σ(wo/coh) (cm²/atom)')
        ws1.cell(column=8, row=1, value='σ(w/coh) (cm²/atom)')

        ws1.cell(column=9, row=1, value='μ/ρ(coh)  (cm²/g)')
        ws1.cell(column=10, row=1, value='μ/ρ(incoh)  (cm²/g)')
        ws1.cell(column=11, row=1, value='μ/ρ(pe)  (cm²/g)')
        ws1.cell(column=12, row=1, value='μ/ρ(pair)  (cm²/g)')
        ws1.cell(column=13, row=1, value='μ/ρ(trip)  (cm²/g)')

        ws1.cell(column=14, row=1, value='μ/ρ(wo/coh) (cm²/g)')
        ws1.cell(column=15, row=1, value='μ/ρ(w/coh) (cm²/g)')
    elif sigma_flag:
        ws1.cell(column=2, row=1, value='σ(coh)  (cm²/atom)')
        ws1.cell(column=3, row=1, value='σ(incoh)  (cm²/atom)')
        ws1.cell(column=4, row=1, value='σ(pe)  (cm²/atom)')
        ws1.cell(column=5, row=1, value='σ(pair)  (cm²/atom)')
        ws1.cell(column=6, row=1, value='σ(trip)  (cm²/atom)')

        ws1.cell(column=7, row=1, value='σ(wo/coh) (cm²/atom)')
        ws1.cell(column=8, row=1, value='σ(w/coh) (cm²/atom)')
        
    else:
        ws1.cell(column=2, row=1, value='μ/ρ(coh)  (cm²/g)')
        ws1.cell(column=3, row=1, value='μ/ρ(incoh)  (cm²/g)')
        ws1.cell(column=4, row=1, value='μ/ρ(pe)  (cm²/g)')
        ws1.cell(column=5, row=1, value='μ/ρ(pair)  (cm²/g)')
        ws1.cell(column=6, row=1, value='μ/ρ(trip)  (cm²/g)')

        ws1.cell(column=7, row=1, value='μ/ρ(wo/coh) (cm²/g)')
        ws1.cell(column=8, row=1, value='μ/ρ(w/coh) (cm²/g)')

    mev=list(array(e_range)/1000000)
    for row in range(1, len(e_range) + 1):
        ws1.cell(column=1, row=row + 1, value=mev[row - 1])
        if both_flag:
            ws1.cell(column=2, row=row + 1, value=to_precision(myu_w_energy[e_range[row - 1]][4],4))
            ws1.cell(column=3, row=row + 1, value=to_precision(myu_w_energy[e_range[row - 1]][5],4))
            ws1.cell(column=4, row=row + 1, value=to_precision(myu_w_energy[e_range[row - 1]][6],4))
            ws1.cell(column=5, row=row + 1, value=to_precision(myu_w_energy[e_range[row - 1]][7],4))
            ws1.cell(column=6, row=row + 1, value=to_precision(myu_w_energy[e_range[row - 1]][8],4))

            ws1.cell(column=7, row=row + 1, value=to_precision(myu_w_energy[e_range[row - 1]][5] + myu_w_energy[e_range[row - 1]][6] + myu_w_energy[e_range[row - 1]][7]+\
                    myu_w_energy[e_range[row - 1]][8],4))
            ws1.cell(column=8, row=row + 1, value=to_precision(myu_w_energy[e_range[row - 1]][5] + myu_w_energy[e_range[row - 1]][6] + myu_w_energy[e_range[row - 1]][7]+\
                    myu_w_energy[e_range[row - 1]][8] + myu_w_energy[e_range[row - 1]][4],4))
            
            
            ws1.cell(column=9, row=row + 1, value=to_precision(myu_w_energy[e_range[row - 1]][9],4))
            ws1.cell(column=10, row=row + 1, value=to_precision(myu_w_energy[e_range[row - 1]][10],4))
            ws1.cell(column=11, row=row + 1, value=to_precision(myu_w_energy[e_range[row - 1]][11],4))
            ws1.cell(column=12, row=row + 1, value=to_precision(myu_w_energy[e_range[row - 1]][12],4))
            ws1.cell(column=13, row=row + 1, value=to_precision(myu_w_energy[e_range[row - 1]][13],4))

            ws1.cell(column=14, row=row + 1, value=to_precision(myu_w_energy[e_range[row - 1]][10] + myu_w_energy[e_range[row - 1]][11] + myu_w_energy[e_range[row - 1]][12]+\
                    myu_w_energy[e_range[row - 1]][13],4))
            ws1.cell(column=15, row=row + 1, value=to_precision(myu_w_energy[e_range[row - 1]][10] + myu_w_energy[e_range[row - 1]][11] + myu_w_energy[e_range[row - 1]][12]+\
                    myu_w_energy[e_range[row - 1]][13] + myu_w_energy[e_range[row - 1]][9],4))
        elif sigma_flag:
            ws1.cell(column=2, row=row + 1, value=to_precision(myu_w_energy[e_range[row - 1]][4],4))
            ws1.cell(column=3, row=row + 1, value=to_precision(myu_w_energy[e_range[row - 1]][5],4))
            ws1.cell(column=4, row=row + 1, value=to_precision(myu_w_energy[e_range[row - 1]][6],4))
            ws1.cell(column=5, row=row + 1, value=to_precision(myu_w_energy[e_range[row - 1]][7],4))
            ws1.cell(column=6, row=row + 1, value=to_precision(myu_w_energy[e_range[row - 1]][8],4))

            ws1.cell(column=7, row=row + 1, value=to_precision(myu_w_energy[e_range[row - 1]][5] + myu_w_energy[e_range[row - 1]][6] + myu_w_energy[e_range[row - 1]][7]+\
                    myu_w_energy[e_range[row - 1]][8],4))
            ws1.cell(column=8, row=row + 1, value=to_precision(myu_w_energy[e_range[row - 1]][5] + myu_w_energy[e_range[row - 1]][6] + myu_w_energy[e_range[row - 1]][7]+\
                    myu_w_energy[e_range[row - 1]][8] + myu_w_energy[e_range[row - 1]][4],4))
        else:
            if len(elements_list) == 1:
                cor = 5
            else:
                cor = 0
            ws1.cell(column=2, row=row + 1, value=to_precision(myu_w_energy[e_range[row - 1]][4+cor],4))
            ws1.cell(column=3, row=row + 1, value=to_precision(myu_w_energy[e_range[row - 1]][5+cor],4))
            ws1.cell(column=4, row=row + 1, value=to_precision(myu_w_energy[e_range[row - 1]][6+cor],4))
            ws1.cell(column=5, row=row + 1, value=to_precision(myu_w_energy[e_range[row - 1]][7+cor],4))
            ws1.cell(column=6, row=row + 1, value=to_precision(myu_w_energy[e_range[row - 1]][8+cor],4))
            
            ws1.cell(column=7, row=row + 1, value=to_precision(myu_w_energy[e_range[row - 1]][5+cor] + myu_w_energy[e_range[row - 1]][6+cor] + myu_w_energy[e_range[row - 1]][7+cor]+\
                    myu_w_energy[e_range[row - 1]][8+cor],4))
            ws1.cell(column=8, row=row + 1, value=to_precision(myu_w_energy[e_range[row - 1]][5+cor] + myu_w_energy[e_range[row - 1]][6+cor] + myu_w_energy[e_range[row - 1]][7+cor]+\
                    myu_w_energy[e_range[row - 1]][8+cor] + myu_w_energy[e_range[row - 1]][4+cor],4))
    ws1.column_dimensions['A'].width = 28
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 18
    ws1.column_dimensions['D'].width = 18
    ws1.column_dimensions['E'].width = 18
    ws1.column_dimensions['F'].width = 18
    ws1.column_dimensions['G'].width = 18
    ws1.column_dimensions['H'].width = 18
    ws1.column_dimensions['I'].width = 18
    ws1.column_dimensions['J'].width = 18
    ws1.column_dimensions['K'].width = 18
    ws1.column_dimensions['L'].width = 18
    ws1.column_dimensions['M'].width = 18
    ws1.column_dimensions['N'].width = 18
    ws1.column_dimensions['O'].width = 18
    
    
    wb.save(filename=dest_filename)
    
    
    

def kerma_1(relative_to_choice='AIR'):
    global elements_list
    global dict_comp
    global mec_comp
    global xray_e_range
    mec_comp = {}
    mec_comp1 = {}

    elements_mec_1 = {}
    for i in elements_list:
        z = element(i).atomic_number
        zin = ''
        if z < 10:
            zin = '0' + str(z)
        else:
            zin = str(z)
        loc = 'XRay_data1/DATA' + zin
        x, y = loadtxt(loc, usecols=(0, 2), unpack=True)
        x11 = list(x)
        y11 = list(y)
        x1 = []
        y1 = []
        for a in x11:
            if a in xray_e_range:
                x1.append(a)
                y1.append(y11[x11.index(a)])

        elements_mec_1[z] = y1

    for e in xray_e_range:
        z = 0.0
        for j in elements_list:
            z = z + (weight_fraction[element(j).atomic_number] * \
                     elements_mec_1[element(j).atomic_number][xray_e_range.index(e)])

        mec_comp1[e] = z

    loc = 'XRay_Comp1/DATA_' + relative_to_choice
    x, y = loadtxt(loc, usecols=(0, 2), unpack=True)
    x1 = list(x)
    y1 = list(y)
    mec_rel = {}
    for num,x in enumerate(x1):
        mec_rel[x] = y1[num]

    denom = 0.0
    sigmaa = {}
    for j in elements_list:
        denom = denom + \
            (weight_fraction[element(j).atomic_number] / element(j).atomic_weight)
    denom1 = denom * n
    for e in xray_e_range:
        sigmaa[e] = mec_comp1[e]/denom1  
#######################################################
    ele_x_int_cross = []

    for i in range(1, 93):
        z = element(i).atomic_number
        zin = ''
        if z < 10:
            zin = '0' + str(z)
        else:
            zin = str(z)
        loc = 'XRay_data1/DATA' + zin
        a=read_csv(loc,delim_whitespace=True,header=None,usecols=[0,2])
        y = array(a[2])
        
        y1 = y / denom1
        ele_x_int_cross.append(list(y1))
    zno = np.arange(1,93)
    params = list(map(list,zip(*ele_x_int_cross)))
    z_comp = []
    z_comp[:] = [element(j).atomic_number for j in elements_list]

    zeff1={}
    
    for e in xray_e_range:
        solutions = InterpolatedUnivariateSpline(zno, params[xray_e_range.index(e)] - sigmaa[e]).roots()
        zeff1[e] = solutions
        if len(elements_list)==1:
            zeff1[e] = [z_comp[0]]
    avg = 0
    Aavg = 0
    z_total = 0
    for j in z_comp:
        avg = avg +(j * weight_fraction[j])
        Aavg = Aavg + ( dict_comp[j] * element(j).mass )
        z_total = z_total + dict_comp[j]
    Aavg = Aavg / z_total
    
    for e in xray_e_range:
        if len(zeff1[e])>1:
            zeff = min(zeff1[e], key=lambda x1:abs(x1-avg))
        else:
            zeff=zeff1[e][0]
#######################################################

        mec_comp[e] = [mec_comp1[e] / mec_rel[e], mec_comp1[e], zeff, mec_comp1[e] / sigmaa[e] * zeff]



def write_mec_to_excel_sheet():
    global mec_comp
    global xray_e_range
    
    sub = formula_for_text

    dest_filename = 'Excel_Sheets/Photon mass-energy absorption coefficients.xlsx'
    if os.path.exists(dest_filename):
        wb = load_workbook(dest_filename)
        if frac_flag:
            sub = sub +'('
            for j in elements_list:
                sub = sub + str(round(weight_fraction[element(j).atomic_number],2)) + ','
            sub = sub.strip(',')
            sub = sub + ')'
            ws1 = wb.create_sheet(str(sub[:27]+'...'), 0)
        else:
            ws1 = wb.create_sheet(str(sub[:27]+'...'), 0)
    else:
        wb = Workbook()
        ws1 = wb.active
        if frac_flag:
            sub = sub +'('
            for j in elements_list:
                sub = sub + str(round(weight_fraction[element(j).atomic_number],2)) + ','
            sub = sub.strip(',')
            sub = sub + ')'
        ws1.title = str(sub[:27]+'...')
    ws1.cell(column=1, row=1, value='Energy (MeV)')

    ws1.cell(column=2, row=1, value='MEC μₑₙ/ρ (cm²/g)')
    if not False and len(elements_list)>1:
        ws1.cell(column=3, row=1, value='Z PEAeff')
        ws1.cell(column=4, row=1, value='N PEAeff (electrons/g)')

    for row in range(1, len(xray_e_range) + 1):
        ws1.cell(column=1, row=row + 1, value=xray_e_range[row - 1])

        ws1.cell(column=2, row=row + 1, value=round(mec_comp[xray_e_range[row - 1]][1],3))
        if not False and len(elements_list)>1:
            ws1.cell(column=3, row=row + 1, value=round(mec_comp[xray_e_range[row - 1]][2],3))
            ws1.cell(column=4, row=row + 1, value=to_precision(mec_comp[xray_e_range[row - 1]][3],4))
    ws1.column_dimensions['A'].width = 13
    ws1.column_dimensions['B'].width = 25
    ws1.column_dimensions['C'].width = 25
    ws1.column_dimensions['D'].width = 25
    
    wb.save(filename=dest_filename)
    
    
    

def write_kerma_to_excel_sheet(choice_r):
    global mec_comp
    global xray_e_range
    
    sub = formula_for_text

    dest_filename = 'Excel_Sheets/Relative KERMA.xlsx'
    if os.path.exists(dest_filename):
        wb = load_workbook(dest_filename)
        if frac_flag:
            sub = sub +'('
            for j in elements_list:
                sub = sub + str(round(weight_fraction[element(j).atomic_number],2)) + ','
            sub = sub.strip(',')
            sub = sub + ')'
            ws1 = wb.create_sheet(str(sub[:27]+'...'), 0)
        else:
            ws1 = wb.create_sheet(str(sub[:27]+'...'), 0)
    else:
        wb = Workbook()
        ws1 = wb.active
        if frac_flag:
            sub = sub +'('
            for j in elements_list:
                sub = sub + str(round(weight_fraction[element(j).atomic_number],2)) + ','
            sub = sub.strip(',')
            sub = sub + ')'
        ws1.title = str(sub[:27]+'...')
    ws1.cell(column=1, row=1, value='Energy (MeV)')
    ws1.cell(column=2, row=1, value=f'{choice_r} KERMA')

    for row in range(1, len(xray_e_range) + 1):
        ws1.cell(column=1, row=row + 1, value=xray_e_range[row - 1])
        ws1.cell(column=2, row=row + 1, value=round(mec_comp[xray_e_range[row - 1]][0],3))
    ws1.column_dimensions['A'].width = 13
    ws1.column_dimensions['B'].width = 17
    ws1.column_dimensions['C'].width = 17
    
    wb.save(filename=dest_filename)
    
    
    

def plot_parameter(energy,parameter,para_name,index=0):
    global formula_for_text
    global weight_fraction
    x=energy
    y=[]
    plt.ylabel('$%s$'%para_name, fontname='Calibri')
    plt.xlabel('$E\ (MeV)$', fontname='Calibri')
    if para_name=='Z_{eq}':
        pass
    plt.ticklabel_format(
        axis='both', style='sci', scilimits=(
            0,0), useMathText=True)
    plt.tick_params(
        axis='both',direction='in',which='both',top=True,right=True
    )
    if isinstance(parameter[x[index]],float): 
        for j in x:
            y.append(parameter[j])
    else:
        for j in x: 
            y.append(parameter[j][index])
    if x[0]>1:
        x=list(array(x)/1000000)
    name=Substance.from_formula(formula_for_text).unicode_name
    if frac_flag:
            sub = name +'('
            for j in elements_list:
                sub = sub + str(round(weight_fraction[element(j).atomic_number],2)) + ','
            sub = sub.strip(',')
            name = sub + ')'
    if para_name in ['Z_{eff}','Z_{eq}','Relative\KERMA','Z_{PEAeff}']:
        plt.semilogx(x, y, '-x', markersize=5, label=name)
    else:
        plt.loglog(x, y, '-x', markersize=5, label=name)
    plt.legend(loc='upper right')
    plt.show()
    plt.close()

def plot_neff(energy,parameter,para_name='N_{eff}\ (electrons/g)'):
    
    global formula_for_text
    x=energy
    y=[]
    plt.ylabel('$%s$'%para_name, fontname='Calibri')
    plt.xlabel('$E\ (MeV)$', fontname='Calibri')
    plt.ticklabel_format(
        axis='both', style='sci', scilimits=(
            0,0), useMathText=True)
    plt.tick_params(
        axis='both',direction='in',which='both',top=True,right=True
    )
        
    if isinstance(parameter[x[0]],float): 
        for j in x:
            y.append(parameter[j])
    else:
        for j in x: 
            y.append(parameter[j][6])

    if x[0]>1:
        x=list(array(x)/1000000)
    name=Substance.from_formula(formula_for_text).unicode_name
    if frac_flag:
            sub = name +'('
            for j in elements_list:
                sub = sub + str(round(weight_fraction[element(j).atomic_number],2)) + ','
            sub = sub.strip(',')
            name = sub + ')'
    plt.semilogx(x, y, '-x', markersize=5, label=name)
    plt.legend(loc='upper right')
    plt.show()
    plt.close()
    


def interpolate_e(energy,parameter,new_energy,num=0):
    y=[]
    
    if isinstance(parameter[energy[0]],float): 
        for j in energy:
            y.append(parameter[j])
    else:
        for j in energy: 
            y.append(parameter[j][num])
    f = interp1d(energy,y,kind='cubic')
    return(f(new_energy))

def interpolate_e_linear(energy,parameter,new_energy,num=0):
    y=[]
    
    if isinstance(parameter[energy[0]],float): 
        for j in energy:
            y.append(parameter[j])
    else:
        for j in energy: 
            y.append(parameter[j][num])
    f = interp1d(energy,y,kind='slinear')
    return(f(new_energy))
def methods():
    global formula_for_text
    global e_range
    global xray_e_range
    global myu_w_energy
    global data
    global proton_energy_range
    global alpha_energy_range
    global B
    global BE
    global mfps
    global do_what
    global energy_flag
    global op
    global icru_mat
    global custom_energies
    single_element_flag = len(elements_list) == 1


    ABCD1 = [

        'Photon interaction parameters ~',
        'Charged particles interaction parameters ~',
        'Exit current computation'
    ]

    ABCD = [
        'Proton interaction',
        'Electron interaction',
        'Alpha particle interaction',
        'Back',
        'Exit current computation'
    ]

    A12345 = [

        'Partial/total interaction cross sections and mass attenuation coefficients',
        'Photon attenuation - Zeff and Neff (electrons/g) ~',
        'Photon energy absorption coefficients (cm²/g), Z PEAeff, N PEAeff (electrons/g)',
        'Relative KERMA',
        'Equivalent atomic number - Zeq',
        'G-P fitting parameters and buildup factors - EABF, EBF',
        'Back',
        'Exit current computation'
    ]
    
    A4ab = [

        'Direct method',
        'Interpolation method',
        'Back',
        'Exit current computation'
    ]
    if frac_flag:
        A4aba = list(A4ab)
        del A4aba[0]

    output_choices = [

        'Write parameter to excel sheet only',
        'Plot Energy vs. Parameter only',
        'Do both'
    ]
    if single_element_flag:
        ABCD1a = list(ABCD1)
        A12345a = list(A12345)
        del A12345a[1]
        del A12345a[3]
        #del ABCD1a[1]

    questions_2 = [
         {

        'type': 'list',
        'name': 'main',
        'message': 'Select: ',
        'choices': ABCD1a if single_element_flag else ABCD1
        },
        {

        'type': 'list',
        'name': 'main',
        'message': 'Select: ',
        'choices': A12345a if single_element_flag else A12345,
        'when': lambda do_what: do_what['main']==ABCD1[0]
        },
        {

        'type': 'list',
        'name': 'main',
        'message': 'Select: ',
        'choices': ABCD,
        'when': lambda do_what: do_what['main']==ABCD1[1]
        },
        {

        'type': 'list',
        'name': 'main',
        'message': 'Choose computation: ',
        'choices': A4aba if frac_flag else A4ab,
        'when': lambda do_what: do_what['main']==A12345[1]
        }

    ]
    
    output_questions = {

        'type': 'list',
        'name': 'output_type',
        'message': 'Choose output: ',
        'choices': output_choices
    }
    energy_options = [
        'Standard energies','Standard + custom energies'
        ]
    user_energy = [

        {
            'type':'list',
            'name':'main',
            'message':'Energy range for parameters: ',
            'choices': energy_options
        }

    ]
    
    
    a = True
    if do_what == A12345[0]:
        old_energy = list(e_range)
        if energy_flag:
            
            energy_q = custom_energies
            energies = energy_q.split()
            energies[:] = [float(x) for x in energies]
            #Interpolation for custom energies - interaction cross sections and attenuation coefficients
            if old_energy[0]>1:
                energies=list(array(energies)*1000000)
            e_range.extend(energies)
            e_range.sort()
            new_parameter_0 = list(interpolate_e(old_energy,myu_w_energy,e_range))
            new_parameter_1 = list(interpolate_e(old_energy,myu_w_energy,e_range,1))
            new_parameter_2 = list(interpolate_e(old_energy,myu_w_energy,e_range,2))
            new_parameter_3 = list(interpolate_e(old_energy,myu_w_energy,e_range,3))
            new_parameter_4 = list(interpolate_e(old_energy,myu_w_energy,e_range,4))
            new_parameter_5 = list(interpolate_e(old_energy,myu_w_energy,e_range,5))
            new_parameter_6 = list(interpolate_e(old_energy,myu_w_energy,e_range,6))
            new_parameter_7 = list(interpolate_e_linear(old_energy,myu_w_energy,e_range,7))
            new_parameter_8 = list(interpolate_e_linear(old_energy,myu_w_energy,e_range,8))
            if len(elements_list)==1:
                new_parameter_9 = list(interpolate_e(old_energy,myu_w_energy,e_range,9))
                new_parameter_10 = list(interpolate_e(old_energy,myu_w_energy,e_range,10))
                new_parameter_11 = list(interpolate_e(old_energy,myu_w_energy,e_range,11))
                new_parameter_12 = list(interpolate_e_linear(old_energy,myu_w_energy,e_range,12))
                new_parameter_13 = list(interpolate_e_linear(old_energy,myu_w_energy,e_range,13))
            
            for num,e in enumerate(e_range):
                try:
                    myu_w_energy[e]
                except KeyError:
                    a = [new_parameter_0[num],new_parameter_1[num],new_parameter_2[num],new_parameter_3[num],
                        new_parameter_4[num],new_parameter_5[num],new_parameter_6[num],new_parameter_7[num],
                            new_parameter_8[num]]
                    if len(elements_list)==1:
                        a.extend([new_parameter_9[num],new_parameter_10[num],new_parameter_11[num],
                                new_parameter_12[num],new_parameter_13[num]])
                    myu_w_energy[e] = a
                        
        #op = prompt(output_questions,style=style_1)['output_type']
        if op == output_choices[1]:
            a = False
            plot_parameter(e_range,myu_w_energy,'\dfrac{\mu}{\\rho}\ \ (cm^{2}/g)')
        elif op == output_choices[2]:
            plot_parameter(e_range,myu_w_energy,'\dfrac{\mu}{\\rho}\ \ (cm^{2}/g)')
        
        while a:
            try:

                write_myu_to_excel_sheet()
                a = False
            except PermissionError:
                eel.excel_alert()
        
        if energy_flag:
            fl = True if e_range[0]>1 else False
            print('\nEnergy(MeV)\t\tMAC(cm²/g)\n')
            for e in energies:
                if fl:
                    e1 = e/1000000
                else:
                    e1 = e
                print("%f\t\t%.3f"%(e1,myu_w_energy[e][0]))
            print('\n')
        
        e_range = list(old_energy)
        
    elif do_what == A4ab[1]:
        #Zeff by interpolation technique
        zeff_by_log()
        old_energy = list(e_range)
        if energy_flag:
            
            energy_q = custom_energies
            energies = energy_q.split()
            energies[:] = [float(x) for x in energies]
            
            if old_energy[0]>1:
                energies=list(array(energies)*1000000)
            e_range.extend(energies)
            e_range.sort()
            new_parameter_0 = list(interpolate_e(old_energy,zeff_log,e_range))
            new_parameter_1 = list(interpolate_e(old_energy,zeff_log,e_range,1))
            new_parameter_2 = list(interpolate_e(old_energy,zeff_log,e_range,2))
            new_parameter_6 = list(interpolate_e(old_energy,zeff_log,e_range,6))
            
            for num,e in enumerate(e_range):
                try:
                    zeff_log[e]
                except KeyError:
                    zeff_log[e] = [
                        new_parameter_0[num],
                        round(new_parameter_1[num]),
                        round(new_parameter_2[num]),
                        'NA',
                        'NA',
                        'NA',
                        new_parameter_6[num]
                    ]

                    
        #op = prompt(output_questions,style=style_1)['output_type']
        if op == output_choices[1]:
            a = False
            plot_parameter(e_range,zeff_log,'Z_{eff}')
            if not False:
                plot_neff(e_range,zeff_log)
        elif op == output_choices[2]:
            plot_parameter(e_range,zeff_log,'Z_{eff}')
            if not False:
                plot_neff(e_range,zeff_log)
    
        while a:
            try:

                write_zeff_to_excel_sheet()
                a = False
            except PermissionError:
                eel.excel_alert()
        if energy_flag:
            fl = True if e_range[0]>1 else False
            print('\nEnergy (MeV)\t\tZeff\n')
            for e in energies:
                if fl:
                    e1 = e/1000000
                else:
                    e1 = e
                print("%f\t\t%.3f"%(e1,zeff_log[e][0]))
            print('\n')
                
        e_range = list(old_energy)
    elif do_what == A12345[4]:
        zeq_by_R()
        old_energy = list(e_range)
        if energy_flag:
            
            energy_q = custom_energies
            energies = energy_q.split()
            energies[:] = [float(x) for x in energies]
            
            if old_energy[0]>1:
                energies=list(array(energies)*1000000)
            e_range.extend(energies)
            e_range.sort()
            new_parameter_0 = list(interpolate_e(old_energy,zeq_R,e_range))
            new_parameter_1 = list(interpolate_e(old_energy,zeq_R,e_range,1))
            new_parameter_2 = list(interpolate_e(old_energy,zeq_R,e_range,2))
            new_parameter_3 = list(interpolate_e(old_energy,zeq_R,e_range,3))
            new_parameter_4 = list(interpolate_e(old_energy,zeq_R,e_range,4))
            new_parameter_5 = list(interpolate_e(old_energy,zeq_R,e_range,5))
            
            
            for num,e in enumerate(e_range):
                try:
                    zeq_R[e]
                except KeyError:
                    zeq_R[e] = [new_parameter_0[num],
                                round(new_parameter_1[num]),
                                round(new_parameter_2[num]),
                                new_parameter_3[num],
                                new_parameter_4[num],
                                new_parameter_5[num]]

        #op = prompt(output_questions,style=style_1)['output_type']
        if op == output_choices[1]:
            a = False
            plot_parameter(e_range,zeq_R,'Z_{eq}')
        elif op == output_choices[2]:
            plot_parameter(e_range,zeq_R,'Z_{eq}')


        
        while a:
            try:

                write_zeq_by_R_to_excel_sheet()
                a = False
            except PermissionError:
                eel.excel_alert()
        
        if energy_flag:
            fl = True if e_range[0]>1 else False
            print('\nEnergy (MeV)\t\tZeq\n')
            for e in energies:
                if fl:
                    e1 = e/1000000
                else:
                    e1 = e
                print("%f\t\t%.3f"%(e1,zeq_R[e][0]))
            print('\n')

        e_range = list(old_energy)

    elif do_what == A12345[5]:
        
        do_what = zeq_by_R(True)
        #op = prompt(output_questions,style=style_1)['output_type']
        if op == output_choices[1]:
            a = False
            for i,m in enumerate(mfps):
                plot_parameter(array(gp_energy_range)*1000000,B,f'EABF\ at\ {m}\ mfp',5+i)
                plot_parameter(array(gp_energy_range)*1000000,BE,f'EBF\ at\ {m}\ mfp',5+i)
        elif op == output_choices[2]:
            for i,m in enumerate(mfps):
                plot_parameter(array(gp_energy_range)*1000000,B,f'EABF\ at\ {m}\ mfp',5+i)
                plot_parameter(array(gp_energy_range)*1000000,BE,f'EBF\ at\ {m}\ mfp',5+i)

        
        while a:
            try:
                write_BF_to_excel_sheet()
                a = False
            except PermissionError:
                eel.excel_alert()


    elif do_what == A4ab[0]:
        #Zeff by direct method
        zeff_by_Ratio()

        old_energy = list(e_range)
        if energy_flag:
            
            energy_q = custom_energies
            energies = energy_q.split()
            energies[:] = [float(x) for x in energies]
            
            if old_energy[0]>1:
                energies=list(array(energies)*1000000)
            e_range.extend(energies)
            e_range.sort()
            new_parameter_0 = list(interpolate_e(old_energy,zeff_Ratio,e_range))
            new_parameter_1 = list(interpolate_e(old_energy,zeff_Ratio,e_range,1))
            new_parameter_2 = list(interpolate_e(old_energy,zeff_Ratio,e_range,2))
            new_parameter_6 = list(interpolate_e(old_energy,zeff_Ratio,e_range,6))
            
            
            for num,e in enumerate(e_range):
                try:
                    zeff_Ratio[e]
                except KeyError:
                    zeff_Ratio[e] = [new_parameter_0[num],
                    new_parameter_1[num],
                    new_parameter_2[num],'NA','NA','NA',
                    new_parameter_6[num]]

        #op = prompt(output_questions,style=style_1)['output_type']
        if op == output_choices[1]:
            a = False
            plot_parameter(e_range,zeff_Ratio,'Z_{eff}')
            plot_neff(e_range,zeff_Ratio,'N_{eff}\ (electrons/g)')
        elif op == output_choices[2]:
            plot_parameter(e_range,zeff_Ratio,'Z_{eff}')
            plot_neff(e_range,zeff_Ratio,'N_{eff}\ (electrons/g)')

        
        while a:
            try:

                write_zeff_by_Ratio_to_excel_sheet()
                a = False
            except PermissionError:
                eel.excel_alert()
        if energy_flag:
            fl = True if e_range[0]>1 else False
            print('\nEnergy (MeV)\t\tZeff\n')
            for e in energies:
                if fl:
                    e1 = e/1000000
                else:
                    e1 = e
                print("%f\t\t%.3f"%(e1,zeff_Ratio[e][0]))
            print('\n')
        
        e_range = list(old_energy)

    elif do_what in [A12345[2],A12345[3]]:
        kerma_flag = False
        if do_what == A12345[3]:
            kerma_flag = True 
        if kerma_flag:
            #rel = prompt(rel_choice,style=style_1)['main']
            rel = icru_mat
            kerma_1(rel)
        else:
            kerma_1()

        old_energy = list(xray_e_range)
        if energy_flag:
            
            energy_q = custom_energies
            energies = energy_q.split()
            energies[:] = [float(x) for x in energies]
            
            if old_energy[0]>1:
                energies=list(array(energies)*1000000)
            xray_e_range.extend(energies)
            xray_e_range.sort()
            new_parameter_0 = list(interpolate_e(old_energy,mec_comp,xray_e_range))
            new_parameter_1 = list(interpolate_e(old_energy,mec_comp,xray_e_range,1))
            new_parameter_2 = list(interpolate_e(old_energy,mec_comp,xray_e_range,2))
            new_parameter_3 = list(interpolate_e(old_energy,mec_comp,xray_e_range,3))
            
            for num,e in enumerate(xray_e_range):
                try:
                    mec_comp[e]
                except KeyError:
                    mec_comp[e] = [

                        new_parameter_0[num],
                        new_parameter_1[num],
                        new_parameter_2[num],
                        new_parameter_3[num]
                    
                    ]
                    
                    
        #op = prompt(output_questions,style=style_1)['output_type']
        if op == output_choices[1]:
            a = False
            if kerma_flag:
                plot_parameter(xray_e_range,mec_comp,'Relative\ KERMA')
            else:
                plot_parameter(xray_e_range,mec_comp,'\mu_{en}/\\rho\ (cm^{2}/g)',1)
                if not False and not len(elements_list) == 1:
                    plot_parameter(xray_e_range,mec_comp,'Z_{PEAeff}',2)
                    plot_parameter(xray_e_range,mec_comp,'N_{PEAeff}\ (electrons/g)',3)
        elif op == output_choices[2]:
            if kerma_flag:
                plot_parameter(xray_e_range,mec_comp,'Relative\ KERMA')
            else:
                plot_parameter(xray_e_range,mec_comp,'\mu_{en}/\\rho\ (cm^{2}/g)',1)
                if not False and not len(elements_list) == 1:
                    plot_parameter(xray_e_range,mec_comp,'Z_{PEAeff}',2)
                    plot_parameter(xray_e_range,mec_comp,'N_{PEAeff}\ (electrons/g)',3)

        
        while a:
            try:
                if not kerma_flag:
                    write_mec_to_excel_sheet()
                else:
                    write_kerma_to_excel_sheet(rel)
                a = False
            except PermissionError:
                eel.excel_alert()
        
        if energy_flag and not kerma_flag:
            fl = True if xray_e_range[0]>1 else False
            print('\nEnergy (MeV)\t\tμen/ρ (cm²/g)\n')
            for e in energies:
                if fl:
                    e1 = e/1000000
                else:
                    e1 = e
                print("%f\t\t%.3f"%(e1,mec_comp[e][1]))
            print('\n')
        if energy_flag and kerma_flag:
            fl = True if xray_e_range[0]>1 else False
            print('\nEnergy (MeV)\t\tKa\n')
            for e in energies:
                if fl:
                    e1 = e/1000000
                else:
                    e1 = e
                print("%f\t\t%.3f"%(e1,mec_comp[e][0]))
            print('\n')
        
        xray_e_range = list(old_energy)

        
    elif do_what == ABCD[1]:
        url1 = 'https://physics.nist.gov/cgi-bin/Star/e_table-ut.pl'
        b = True
        with requests.Session() as c:
            while b:
                try:
                    c.post(url1)
                    answer11 = True
                    b = False
                except requests.exceptions.ConnectionError:
                    answer11 = eel.excel_confirm('Internet connection cannot be established and is necessary.\nContinue now?')()
                    b = answer11
        a = False
        if answer11:
            zeff_electron_interaction()
            a = True
            x = list(data[0])
            old_energy = list(x)
            if energy_flag and answer11:
                
                energy_q = custom_energies
                energies = energy_q.split()
                energies[:] = [float(x) for x in energies]
                
                if old_energy[0]>1:
                    energies=list(array(energies)*1000000)
                x.extend(energies)
                x.sort()
                new_parameter_0 = list(interpolate_e(old_energy,zeff_ele,x))
                new_parameter_1 = list(interpolate_e(old_energy,zeff_ele,x,1))
                new_parameter_2 = list(interpolate_e(old_energy,zeff_ele,x,2))
                new_parameter_6 = list(interpolate_e(old_energy,zeff_ele,x,6))
                new_parameter_7 = list(interpolate_e(old_energy,electron_msp,x))
                new_parameter_sc = list(interpolate_e(old_energy,zeff_ele,x,3))
                
                for num,e in enumerate(x):
                    try:
                        zeff_ele[e]
                    except KeyError:
                        zeff_ele[e] = [
                            new_parameter_0[num],
                            round(new_parameter_1[num]),
                            round(new_parameter_2[num]),
                            new_parameter_sc[num],
                            'NA',
                            'NA',
                            new_parameter_6[num]
                            ]
                        electron_msp[e] = round(new_parameter_7[num],3)
                
            #op = prompt(output_questions,style=style_1)['output_type']
            if op == output_choices[1]:
                a = False
                if not single_element_flag:
                    plot_parameter(x,zeff_ele,'Z_{eff}')
                    plot_neff(x,zeff_ele)
                plot_parameter(x,electron_msp,'S(E)/\\rho\ (MeV\ cm^{2}/g)')
            elif op == output_choices[2]:
                if not single_element_flag:
                    plot_parameter(x,zeff_ele,'Z_{eff}')
                    plot_neff(x,zeff_ele)
                plot_parameter(x,electron_msp,'S(E)/\\rho\ (MeV\ cm^{2}/g)')
        
        while a:
            try:
                write_zeff_ele_to_excel_sheet(x)
                a = False
            except PermissionError:
                eel.excel_alert()
        if energy_flag and answer11 and not single_element_flag:
            fl = True if x[0]>1 else False
            print('\nEnergy (MeV)\t\tZeff\t\tS(E)/ρ (cm²/g)\n')
            for e in energies:
                if fl:
                    e1 = e/1000000
                else:
                    e1 = e
                print("%f\t\t%.3f\t\t%.3f"%(e1,zeff_ele[e][0],electron_msp[e]))
            print('\n')
        if frac_flag and answer11:
            x = list(old_energy)

    elif do_what == ABCD[0]:
        #start = time.time()
        zeff_proton_interaction()
        #print(f'Time taken = {start-time.time()}')
        x = proton_energy_range

        old_energy = list(x)
        if energy_flag:
            
            energy_q = custom_energies
            energies = energy_q.split()
            energies[:] = [float(x) for x in energies]
            
            if old_energy[0]>1:
                energies=list(array(energies)*1000000)
            x.extend(energies)
            x.sort()
            new_parameter_0 = list(interpolate_e(old_energy,zeff_proton,x))
            new_parameter_1 = list(interpolate_e(old_energy,zeff_proton,x,1))
            new_parameter_2 = list(interpolate_e(old_energy,zeff_proton,x,2))
            new_parameter_6 = list(interpolate_e(old_energy,zeff_proton,x,6))
            new_parameter_7 = list(interpolate_e(old_energy,proton_msp,x))
            new_parameter_sc = list(interpolate_e(old_energy,zeff_proton,x,3))
            for num,e in enumerate(x):
                try:
                    zeff_proton[e]
                except KeyError:
                    zeff_proton[e] = [
                        new_parameter_0[num],
                        round(new_parameter_1[num]),
                        round(new_parameter_2[num]),
                        new_parameter_sc[num],
                        'NA',
                        'NA',
                        new_parameter_6[num]]
                    proton_msp[e] = round(new_parameter_7[num],3)

        #op = prompt(output_questions,style=style_1)['output_type']
        if op == output_choices[1]:
            a = False
            if not single_element_flag:
                plot_parameter(x,zeff_proton,'Z_{eff}')
                plot_neff(x,zeff_proton)
            plot_parameter(x,proton_msp,'S(E)/\\rho\ (MeV\ cm^{2}/g)')
        elif op == output_choices[2]:
            if not single_element_flag:
                plot_parameter(x,zeff_proton,'Z_{eff}')
                plot_neff(x,zeff_proton)
            plot_parameter(x,proton_msp,'S(E)/\\rho\ (MeV\ cm^{2}/g)')
        while a:
            try:

                write_zeff_proton_to_excel_sheet(x)
                a = False
            except PermissionError:
                eel.excel_alert()
        
        if energy_flag and not single_element_flag:
            fl = True if x[0]>1 else False
            print('\nEnergy (MeV)\t\tZeff\t\tS(E)/ρ (cm²/g)\n')
            for e in energies:
                if fl:
                    e1 = e/1000000
                else:
                    e1 = e
                print("%f\t\t%.3f\t\t%.3f"%(e1,zeff_proton[e][0],proton_msp[e]))
            print('\n')

        x = list(old_energy)

    elif do_what == ABCD[2]:
        zeff_alpha_interaction()
        x = alpha_energy_range

        old_energy = list(x)
        if energy_flag:
            
            energy_q = custom_energies
            energies = energy_q.split()
            energies[:] = [float(x) for x in energies]
            
            if old_energy[0]>1:
                energies=list(array(energies)*1000000)
            x.extend(energies)
            x.sort()
            new_parameter_0 = list(interpolate_e(old_energy,zeff_alpha,x))
            new_parameter_1 = list(interpolate_e(old_energy,zeff_alpha,x,1))
            new_parameter_2 = list(interpolate_e(old_energy,zeff_alpha,x,2))
            new_parameter_6 = list(interpolate_e(old_energy,zeff_alpha,x,6))
            new_parameter_7 = list(interpolate_e(old_energy,alpha_msp,x))
            new_parameter_sc = list(interpolate_e(old_energy,zeff_alpha,x,3))
            for num,e in enumerate(x):
                try:
                    zeff_alpha[e]
                except KeyError:
                    zeff_alpha[e] = [
                        new_parameter_0[num],
                        round(new_parameter_1[num]),
                        round(new_parameter_2[num]),
                        new_parameter_sc[num],
                        'NA',
                        'NA',
                        new_parameter_6[num]
                        ]
                    alpha_msp[e] = round(new_parameter_7[num],3)

        #op = prompt(output_questions,style=style_1)['output_type']
        if op == output_choices[1]:
            a = False
            if not single_element_flag:
                plot_parameter(x,zeff_alpha,'Z_{eff}')
                plot_neff(x,zeff_alpha)
            plot_parameter(x,alpha_msp,'S(E)/\\rho\ (MeV\ cm^{2}/g)')
        elif op == output_choices[2]:
            if not single_element_flag:
                plot_parameter(x,zeff_alpha,'Z_{eff}')
                plot_neff(x,zeff_alpha)
            plot_parameter(x,alpha_msp,'S(E)/\\rho\ (MeV\ cm^{2}/g)')
        while a:
            try:

                write_zeff_alpha_to_excel_sheet(x)
                a = False
            except PermissionError:
                eel.excel_alert()
        
        if energy_flag and not single_element_flag:
            fl = True if x[0]>1 else False
            print('\nEnergy (MeV)\t\tZeff\t\tS(E)/ρ (cm²/g)\n')
            for e in energies:
                if fl:
                    e1 = e/1000000
                else:
                    e1 = e
                print("%f\t\t%.3f\t\t%.3f"%(e1,zeff_alpha[e][0],alpha_msp[e]))
            print('\n')

        x = list(old_energy)


def CreateFolder(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except OSError:
        print('Error: Creating directory. ' + directory)

def CreateLog(dat1, loc = "InputLog.log"):
    with open(loc,"a") as text_file:
        text_file.write(json.dumps(dat1))
        text_file.write('\n')

CreateFolder('Excel_Sheets')

@eel.expose
def main1( comp_0a='H', do_what_now='Back', output='Do both', ff1=False,comp_1a='H', comp_2a='1', eflag=False,  mfp='1', density='1', \
    rel_mat='Air', custom_energies_list=''):
    #print("Entered.")
    global frac_flag
    global comp_0
    global comp_1
    global comp_2
    global energy_flag
    global do_what 
    global op
    global mean_free_path
    global density_mat
    global icru_mat
    global custom_energies
    custom_energies = custom_energies_list
    icru_mat = rel_mat
    density_mat = density
    mean_free_path = mfp
    op = output
    do_what = do_what_now
    energy_flag = eflag
    comp_0 = comp_0a
    comp_1 = comp_1a
    comp_2 = comp_2a
    frac_flag = ff1
    input_log = {}
    now = datetime.now()
    dt_string = now.strftime("%d/%m/%Y %H:%M:%S")
    input_log["Log time"] = dt_string
    input_log["Composition_0"] = comp_0
    input_log["Composition_1a"] = comp_1
    input_log["Composition_1b"] = comp_2
    input_log["Composition_1b"] = comp_2
    input_log["Parameter"] = do_what
    input_log["Output"] = op
    input_log["Energies"] = custom_energies
    input_log["KERMA Material"] = icru_mat
    input_log["Density"] = density_mat
    input_log["MFP"] = mean_free_path
    
    CreateLog(input_log)
    start_time = time.process_time()
    # print("Here")
    fetch_compound()
    myu()
    methods()
    CreateLog(f'Time elapsed: {time.process_time() - start_time}s')
    eel.excel_alert("Computation complete!")

    # print("Exit.")
eel.init('web')
eel.start('landing2.4.html',size=(1024, 550))
# eel.start('index.html', size=(980, 660))